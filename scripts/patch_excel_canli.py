# -*- coding: utf-8 -*-
"""Add or replace 'Canli_AltYapi' in gbsoft-fiyat-modeli.xlsx with live infrastructure lines."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

PATH = r"c:\Users\satog\OneDrive\Desktop\dengizek\gbsoft-fiyat-modeli.xlsx"
SHEET = "Canli_AltYapi"

FILL_SEC = PatternFill("solid", fgColor="D9E1F2")
FILL_HDR = PatternFill("solid", fgColor="1F3A5F")
FONT_HDR = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, len(wb.sheetnames))
    w = 72
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 52

    ws.merge_cells("A1:D1")
    a1 = ws["A1"]
    a1.value = (
        "Canlı altyapı (üretim) — Geliştirme araçları (IDE, agent abonelikleri) hariç"
    )
    a1.font = Font(bold=True, size=12)

    ws.merge_cells("A2:D2")
    a2 = ws["A2"]
    a2.value = (
        "Barındırma (Vercel), veritabanı (Supabase/Postgres), video (Cloudflare Stream), "
        "depolama (R2), e-posta, hata/ürün analitiği, cache (Redis), alan adı, opsiyonel CF Pro. "
        "Yapay zekâ API (LLM/STT/TTS) ayrı kalem: kullanım başı — ayrıntı AI_Uretim sayfası."
    )
    a2.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 48

    r = 4
    ws.cell(r, 1, "PARAMETRELER")
    ws.cell(r, 1).font = BOLD
    for c in (1, 2, 3, 4):
        ws.cell(r, c).fill = FILL_SEC
    r += 1
    # B5=kur, B6=katalog dk, B7=delivery dk
    ws.cell(r, 1, "USD/TL kuru (Ozet B5 referansı)")
    ws.cell(r, 2, "=Ozet!$B$5")
    ws.cell(r, 3, None)
    ws.cell(
        r,
        4,
        "Kur; TL sütunları = USD × bu hücre. Nisan 2026 tahmini Ozet’te 40 olabilir.",
    )
    r += 1
    ws.cell(r, 1, "Katalog toplam video (dk) — Stream saklama hacmi")
    ws.cell(r, 2, 2500)
    ws.cell(r, 4, "Tüm turların depodaki toplam dakikası. Cloudflare: 5 USD / 1000 dk saklama")
    r += 1
    ws.cell(r, 1, "Aylık toplam izlenen video (dk) — Stream delivery")
    ws.cell(r, 2, 50000)
    ws.cell(
        r,
        4,
        "Aylık oynatılan toplam dakika. Cloudflare: 1 USD / 1000 dk — ana değişken gider",
    )
    r += 1

    r = 9
    for c, h in enumerate(
        ("Kalem", "USD/ay", "TL/ay", "Notlar (kaynak: sağlayıcı fiyat sayfaları, 2025–2026)"),
        1,
    ):
        cell = ws.cell(r, c, h)
        cell.fill = FILL_HDR
        cell.font = FONT_HDR
    r += 1

    lines: list[tuple[str, str | float, str]] = [
        (
            "Vercel Pro — platform + 1 koltuk (taban)",
            20,
            "vercel.com/pricing — ayrıca kullanım kredisi/edge; trafik arttıkça artar",
        ),
        (
            "Vercel — edge/bandwidth/görüntü (tahmini ek)",
            30,
            "Düşük-orta trafik için tahmini; yüksek trafikte 50–200+ olabilir",
        ),
        (
            "Supabase Pro — taban (Postgres, Auth, Storage kotaları)",
            25,
            "supabase.com/pricing",
        ),
        (
            "Supabase — compute + MAU/egress aşımı (tahmini)",
            40,
            "Büyüyünce; tek başına 25–150 USD bandı tipik (MAU, egress, large compute)",
        ),
        (
            "Cloudflare R2 — nesne depolama (görsel, export, yedek)",
            8,
            "0,015 USD/GB-ay; egress yok — developers.cloudflare.com/r2/pricing",
        ),
        (
            "Cloudflare Stream — video saklama (formül)",
            "=ROUND($B$6/1000*5,2)",
            "5 USD / 1000 dakika; B6 = katalog toplam dk",
        ),
        (
            "Cloudflare Stream — video oynatma / delivery (formül)",
            "=ROUND($B$7/1000*1,2)",
            "1 USD / 1000 izlenen dakika; B7 = aylık delivery dk",
        ),
        (
            "Cloudflare (Pro) — WAF/özel ihtiyaç (opsiyonel)",
            0,
            "Gerekirse; Stream’den ayrı — cloudflare.com/plans",
        ),
        (
            "Resend — transactional e-posta",
            0,
            "Kota içi ücretsiz; 50k+ e-posta için Pro ~20 USD — resend.com/pricing",
        ),
        (
            "Sentry — hata izleme (Team, tipik)",
            26,
            "Yıllık faturalamada ~26; aylık farklı — sentry.io/pricing; ücretsiz Developer alternatif",
        ),
        (
            "PostHog — ürün analitiği",
            0,
            "1M event/ay ücretsiz; üstü pay-as-you-go — posthog.com/pricing",
        ),
        (
            "Upstash — Redis (cache, rate limit, kuyruk)",
            10,
            "Sabit 10+ veya PAYG; upstash.com/pricing",
        ),
        (
            "Edge Functions / arka plan (Supabase, CF Workers) aşım (tahmini)",
            15,
            "İş yüküne göre; düşük başlangıçta 0–15 USD",
        ),
        (
            "Alan adı + DNS (ortalama/ay)",
            1.5,
            "Yıllık ~18 USD / 12",
        ),
        (
            "Yedekleme / PITR eklentisi (Supabase, opsiyonel)",
            0,
            "Gerekirse ayrı satır; Supabase fiyat listesi",
        ),
        (
            "Beklenmeyen transfer / e-posta aşımı (tampon)",
            10,
            "Egress veya sürpriz kalemler için küçük tampon",
        ),
    ]

    first_data = r
    for label, usd, note in lines:
        ws.cell(r, 1, label)
        if isinstance(usd, str) and usd.startswith("="):
            ws.cell(r, 2, usd)
        else:
            ws.cell(r, 2, usd)
        ws.cell(r, 3, f"=B{r}*$B$5")
        ws.cell(r, 4, note)
        r += 1
    last_data = r - 1

    ws.cell(r, 1, "TOPLAM (altyapı + video formülleri)")
    ws.cell(r, 1).font = BOLD
    ws.cell(r, 2, f"=SUM(B{first_data}:B{last_data})")
    ws.cell(r, 3, f"=SUM(C{first_data}:C{last_data})")
    ws.cell(r, 4, "B6/B7 değiştikçe Stream satırları ve toplam güncellenir")
    for c in (1, 2, 3, 4):
        ws.cell(r, c).fill = FILL_SEC

    r += 2
    ws.merge_cells(f"A{r}:D{r}")
    ai = ws.cell(r, 1)
    ai.value = (
        "Yapay zekâ API (OpenAI, Anthropic, STT, TTS, embedding): Bu sayfaya dahil DEĞİL. "
        "Aylık toplam = kullanım × birim fiyat; tahmin için AI_Uretim sayfasındaki parametreler."
    )
    ai.alignment = Alignment(wrap_text=True, vertical="top")
    ai.font = Font(italic=True)
    ws.row_dimensions[r].height = 36

    wb.save(PATH)
    print(f"OK: {SHEET} eklendi — {PATH}")


if __name__ == "__main__":
    main()
