# -*- coding: utf-8 -*-
"""Apply sourced USD pricing notes + refreshed infra USD rows (April 2026 refs)."""
import openpyxl
from openpyxl.styles import Alignment

PATH = r"c:\Users\satog\OneDrive\Desktop\dengizek\gbsoft-fiyat-modeli.xlsx"


def main() -> None:
    wb = openpyxl.load_workbook(PATH)

    srv = wb["Sunucu_Barindirma"]
    # E列 USD； G列 Notlar (col 7). Rows match sheet layout:
    updates_srv = {
        5: (
            24,
            "digitalocean.com/pricing — Basic 2vCPU/4GB $24/ay (dev); üretim 4vCPU/8GB $48/ay; Docker Compose: app+postgres+redis+nginx",
        ),
        6: (
            0,
            "VPS içindeki Docker container — postgres:16-alpine; veri /volumes/postgres; pg_dump→R2 backup günlük cron",
        ),
        8: (
            8,
            "developers.cloudflare.com/r2/pricing — Standard $0.015/GB-ay (örnek ~530 GB≈$8); pg_dump backup için de kullanılır",
        ),
        9: (
            5,
            "bunny.net/stream — encoding ÜCRETSİZ; depolama $0.01/GB-ay; delivery $0.005/GB (Avrupa/TR); signed token auth dahil",
        ),
        17: (
            25,
            "openai.com/api/pricing — Whisper $0.006/dk; TTS karakter başına (API PAYG tahmini)",
        ),
    }
    for row, (usd, note) in updates_srv.items():
        srv.cell(row, 5, usd)
        srv.cell(row, 7, note)

    srv["E6"].number_format = "0"
    srv["D6"].value = "VPS içinde (ayrı ücret yok)"
    srv["B6"].value = "Postgres (self-hosted, Docker)"
    srv["C6"].value = "DigitalOcean VPS"
    srv["A5"].value = "VPS + App + DB + Cache"
    srv["B5"].value = "Docker (Next.js + Postgres + Redis + Nginx)"
    srv["C5"].value = "DigitalOcean"
    srv["D5"].value = "$24 Droplet (dev) / $48 üretim"
    srv["A6"].value = "Database"
    srv["A7"].value = "Cache/Queue"
    srv["B7"].value = "Redis (self-hosted, Docker)"
    srv["C7"].value = "DigitalOcean VPS"
    srv["D7"].value = "VPS içinde (ayrı ücret yok)"
    srv["E7"].value = 0

    gel = wb["Geliştirme_Araclari"]
    gel_notes = {
        5: "claude.com/pricing/max — Max 5x $100/kişi/ay (web)",
        6: "cursor.com/pricing — Pro $20/ay",
        8: "console.anthropic.com — API Sonnet $3/$15 per MTok (docs.anthropic.com)",
        9: "platform.openai.com — GPT/API PAYG tahmini",
    }
    for row, note in gel_notes.items():
        gel.cell(row, 7, note)

    ai = wb["AI_Uretim"]
    ai.cell(10, 3).value = (
        "openai.com/api/pricing — TTS $15/1M karakter (Std); dakika başına tahmini yaklaşım"
    )
    ai.cell(9, 3).value = "openai.com/api/pricing — Whisper $0.006/dk"

    oz = wb["Ozet"]
    # Kaynak bloğu — özet altına (satır kaydırmadan)
    oz.merge_cells("A35:G39")
    c = oz["A35"]
    c.value = (
        "Fiyat kaynakları (güncel resmi — kontrol tarihi Nisan 2026):\n"
        "• DigitalOcean Droplet — digitalocean.com/pricing ($24/ay dev, $48/ay üretim); Docker Compose ile app+postgres+redis.\n"
        "• Postgres self-hosted — VPS içinde Docker container; ayrı ücret yok; pg_dump→R2 backup.\n"
        "• Bunny Stream — bunny.net/stream (encoding ücretsiz; $0.01/GB-ay depolama; $0.005/GB delivery Avrupa/TR).\n"
        "• Cloudflare R2 — developers.cloudflare.com/r2/pricing ($0.015/GB-ay).\n"
        "• Claude Max — claude.com/pricing/max (Max 5x $100/ay).\n"
        "• Cursor Pro — cursor.com/pricing ($20/ay).\n"
        "• Anthropic Claude API — docs.anthropic.com/pricing (örn. Sonnet $3/$15 per MTok).\n"
        "• OpenAI API — openai.com/api/pricing (Whisper $0.006/dk, ücretler modele göre).\n"
        "Özet USD satırları diğer sayfa referanslarından yeniden hesaplanır; TL = USD × Ozet!B5 kur tahmini."
    )
    c.alignment = Alignment(wrap_text=True, vertical="top")
    oz.row_dimensions[35].height = 140

    wb.save(PATH)
    print("OK:", PATH)


if __name__ == "__main__":
    main()
